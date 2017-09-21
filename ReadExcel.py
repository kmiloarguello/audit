from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell

wb = load_workbook(filename = 'xlsx/BMW_Sales_Standards_2016_ME.xlsx', data_only=True)
sheets = wb.sheetnames[11:12] #Current Sheet

# Arrays
numberCategory = []
index_number_categories = []
rowData = []
cleanedRowData = []

for sheet in sheets: #In current sheet give me the rows and columns
  ws = wb[sheet] # Pass the info as ws variable
  for row in ws.rows: #Get all rows
    numberCategory.insert(0,row[23].value)  #Get the category number and insert the value of row #1 in excel => B
    number_categories_without_filter = next(i for i in numberCategory if i is not None) #Clean of every None and replacing for the previous valid element
    index_number_categories.extend([number_categories_without_filter]) # Store data row into a Array to see its index
    if(row[23].value == "N"): #If X1 to Xn something has N get that result
      n_result = [] # Empty array to add values with N
      for i in range(len(index_number_categories)): # Go through the array
        if index_number_categories[i] == 'N': # set the i into the array if are equal to "N"
          n_result.append(i) # add this index to n_result the N values
      print n_result
      # for column in ws.columns: #Get info of current row that has N in that column
        # n_result = index_number_categories.index("N") #Find all data with value N in that column
    #     rowData.insert(0,column[n_result].value)  #Get the category number and insert the value of row #1 in excel => B
    #     if(len(rowData) == 32): #Get the last item of iterator array
    #       print rowData[14] #Print final data filtered
          # wbp = Workbook(write_only=True) #Call again Workbook with parameter of writing new file
          # wsp = wbp.create_sheet() #Create sheet
          # for i in range(len(rowData)): # In rowData get all the data
          #   wsp.append([rowData[i]]) # append it into a cells
          # wbp.save('result.xlsx') # Save excel file

