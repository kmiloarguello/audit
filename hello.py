# Tkinter lib to create user interface
import sys
from Tkinter import *
from tkFileDialog import askopenfilename
from tkintertable import TableCanvas, TableModel
import tkMessageBox


# Openpyxl libs
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell

numberCategory = []
zerovalue = []
index_number_categories = []
rowData = []
cleanedRowData = []
auditvalue = []
essential = []
standard = []
number = []
requirement = []
comments = []
question = []
observation = []
suggested = []

wb = load_workbook(filename = 'xlsx/BMW_Sales_Standards_2016_ME.xlsx', data_only=True)

sheets = wb.sheetnames[3:12]

myHoja = []
myStandard = []
myNumber = []
myRequeriment = []
myComment = []
myAudit = []
myEssentials = []
myAuditQuestion = []
myObservation = []
mySuggested = []
myN = []
myZero = []
myAComments = []
myPic = []
auditcomments = []
picture = []


for sheet in sheets:
  ws = wb[sheet]

  for row in ws.rows: 
    numberCategory.insert(0,row[23].value)  
    number_categories_without_filter = next(i for i in numberCategory if i is not None)
    index_number_categories.extend([number_categories_without_filter])

    zerovalue.insert(0,row[25].value)  
    zero_categories_without_filter = next(i for i in zerovalue if i is not None)

    auditvalue.insert(0,row[13].value)  
    audit_categories_without_filter = next(i for i in auditvalue if i is not None)

    essential.insert(0,row[15].value)  
    essential_without_filter = next(i for i in essential if i is not None)

    standard.insert(0,row[0].value)  
    standard_categories_without_filter = next(i for i in standard if i is not None)

    number.insert(0,row[1].value)  
    number_categories_without_filter = next(i for i in number if i is not None)

    requirement.insert(0,row[2].value)  
    requirement_categories_without_filter = next(i for i in requirement if i is not None)

    comments.insert(0,row[4].value)  
    comments_categories_without_filter = next(i for i in comments if i is not None)

    question.insert(0,row[17].value)  
    question_categories_without_filter = next(i for i in question if i is not None)

    observation.insert(0,row[19].value)  
    observation_categories_without_filter = next(i for i in observation if i is not None)

    suggested.insert(0,row[21].value)  
    suggested_categories_without_filter = next(i for i in suggested if i is not None)

    auditcomments.insert(0,row[30].value)  
    auditcomments_categories_without_filter = next(i for i in auditcomments if i is not None)

    picture.insert(0,row[30].value)  
    picture_categories_without_filter = next(i for i in picture if i is not None)

    final_audit = audit_categories_without_filter.encode('ascii','ignore')

    if(row[23].value == "N" and zero_categories_without_filter == 0 and 'Audit' in final_audit ):
      myHoja.extend([sheet])
      myN.extend([str(row[23].value)])
      myZero.extend([str(zero_categories_without_filter)])
      myAudit.extend([str(audit_categories_without_filter)])
      myEssentials.extend([str(essential_without_filter)])
      
      print myAudit 
      myStandard.extend([str(standard_categories_without_filter)])
      myNumber.extend([str(number_categories_without_filter)])
      myRequeriment.extend([requirement_categories_without_filter.encode('utf-8')])
      myComment.extend([str(comments_categories_without_filter)])
      myAuditQuestion.extend([str(question_categories_without_filter)])
      myObservation.extend([str(observation_categories_without_filter)])
      mySuggested.extend([str(suggested_categories_without_filter)])
      myAComments.extend([str(auditcomments_categories_without_filter)])
      myPic.extend([str(picture_categories_without_filter)])
