from Tkinter import *
from PIL import Image, ImageTk
from tkFileDialog import askopenfilename, asksaveasfilename
import tkMessageBox
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
import types
# from pandastable import Table, TableModel
from tkintertable.Tables import TableCanvas
from tkintertable.TableModels import TableModel


myColor = 'white'

class OtherFrame(Toplevel):
    """Main Interaction of the App. Here the app shows the Excel and let to the user upload an image."""
    #----------------------------------------------------------------------
    def __init__(self, original):
      """Constructor"""
      self.original_frame = original
      Toplevel.__init__(self)
      # self.wm_iconbitmap('kapta_mex.ico')
      
      #Center the window

      myW = 1200
      myH = 600

      myWs = root.winfo_screenwidth()
      myHs = root.winfo_screenheight()

      myX = (myWs/2) - (myW/2)
      myY = (myHs/2) - (myH/2)

      self.geometry('%dx%d+%d+%d' % (myW, myH, myX, myY))

      """
      Define Name and background color for the window
      """

      # self.geometry("800x400")
      self.title("K@PTA Auditorias")
      # self.configure(background=myColor)

      """
      Load Menu and Excel for star the app
      After that render the Excel into a Frame.
      """

      self.loadMenu()
      
      self.frame = Frame(self)
      self.loadExcel()
      self.renderExcel()
      self.frame.pack(fill=X)

      """
      Frame to upload the image and manage it.
      """

      self.line = Frame(self, relief=RAISED, borderwidth=1)
      self.line.pack(fill=X,)

      self.image_container = Frame(self)
      Label(self.image_container, text='Imagenes para guardar en excel', font=("Helvetica", 15), foreground='#000').pack(pady=20)
      self.buttonImage()
      self.image_container.pack(fill=X, pady=5)

      """
      Button of CLOSE
      """

      self.frame2 = Frame(self)
      btn = Button(self.frame2, text="Close", command=self.onClose)
      btn.pack()
      self.frame2.pack()

      """
      # Scrollbar funcionality

      self.scrollbar = Scrollbar(self)
      self.scrollbar.pack(side=RIGHT, fill=Y)

      self.listbox = Listbox(self.image_container, yscrollcommand=self.scrollbar.set)
      for i in range(1000):
        self.listbox.insert(END, str(i))
      self.listbox.pack(side=LEFT, fill=BOTH)

      self.scrollbar.config(command=self.listbox.yview)

      """

      self.footer()

    #----------------------------------------------------------------------
    def loadMenu(self):
      self.menu = Menu(self)
      self.config(menu=self.menu)

      self.subMenu = Menu(self.menu,tearoff=0,bg='white')
      self.menu.add_cascade(label='Archivo', menu=self.subMenu)
      self.subMenu.add_command(label='Abrir excel', command="openExcel")
      self.subMenu.add_command(label='Abrir imagen', command="openImage")
      self.subMenu.add_command(label='Exportar excel', command="")
      self.subMenu.add_separator()
      self.subMenu.add_command(label='Acerca de K@PTA', command=self.acercaDe)
      self.subMenu.add_command(label='Salir', command="exitApp")

      self.editMenu = Menu(self.menu,tearoff=0,bg='white')
      self.menu.add_cascade(label='Editar', menu=self.editMenu)
      self.editMenu.add_command(label='Deshacer', command="myFunction")

      self.helpMenu = Menu(self.menu,tearoff=0,bg='white')
      self.menu.add_cascade(label='Ayuda', menu=self.helpMenu)

    #----------------------------------------------------------------------
    """
    LOAD EXCEL Funcionality
    
    1. arraysInit() -> is the initialization of every array
    2. loadExcel() -> Load the excel file
    3. withoutFilter() -> Return a list with cell without None values
    4. loadWorkbook() -> Load from the file, the sheets and each column. Make search functionality
    5. renderExcel() -> Puth the info of file inside of a canvas using tkintertables

    """
    def arraysInit(self):
      self.numberCategory = []
      self.zerovalue = []
      self.index_number_categories = []
      self.rowData = []
      self.cleanedRowData = []
      self.auditvalue = []
      self.essential = []
      self.standard = []
      self.number = []
      self.requirement = []
      self.comments = []
      self.question = []
      self.observation = []
      self.suggested = []
      self.myHoja = []
      self.myStandard = []
      self.myNumber = []
      self.myRequeriment = []
      self.myComment = []
      self.myAudit = []
      self.myEssentials = []
      self.myAuditQuestion = []
      self.myObservation = []
      self.mySuggested = []
      self.myN = []
      self.myZero = []
      self.myAComments = []
      self.myPic = []
      self.auditcomments = []
      self.picture = []        
    
    def loadExcel(self):
      self.filename = askopenfilename( initialdir = "/KAPTA Camilo/python/xlsx",title = "Subir archivo de excel", filetypes = (("Excel Auditorias", ".xlsx"), ("Todos los archivos", "*.*")))  
      self.arraysInit()
      self.loadWorkbook()

    def withoutFilter(self,item,cell):
      """
        The NumberTypes variable stores all numeric data type from types library
        The main key is that if the type of variable is not numeric, encodes it into a  ascii
      """
      NumberTypes = (types.IntType, types.LongType, types.FloatType, types.ComplexType)
      item.insert(0,cell.value)  
      without_filter = next(i for i in item if i is not None)
      if not isinstance(without_filter, NumberTypes):
        return without_filter.encode('ascii', 'ignore')
      else:
        return without_filter

    def loadWorkbook(self):
      """
       Excel Funcionality to make an average of N and Audit Values and return it
       First Load the excel file
      """
      self.wb = load_workbook(filename=self.filename,data_only=True)
      self.sheets = self.wb.sheetnames[3:12]

      for sheet in self.sheets:
        self.ws = self.wb[sheet]
        for row in self.ws.rows:
          final_numberCat = self.withoutFilter(self.numberCategory,row[23])
          final_zero = self.withoutFilter(self.zerovalue,row[25])
          final_audit = self.withoutFilter(self.auditvalue,row[13])
          final_essential = self.withoutFilter(self.essential,row[15])
          final_standard = self.withoutFilter(self.standard,row[0])
          final_number = self.withoutFilter(self.number,row[1])
          final_requirement = self.withoutFilter(self.requirement,row[2])
          final_comments = self.withoutFilter(self.comments,row[4])
          final_question = self.withoutFilter(self.question,row[17])
          final_observation = self.withoutFilter(self.observation,row[19])
          final_suggested = self.withoutFilter(self.suggested,row[21])
          final_auditcomments = self.withoutFilter(self.auditcomments,row[30])
          final_picture = self.withoutFilter(self.picture,row[30])

          
          if(row[23].value == "N" and final_zero == 0 and 'Audit' in final_audit ):
            self.myHoja.extend([sheet])
            self.myN.extend([str(row[23].value)])
            self.myZero.extend([str(final_zero)])
            self.myAudit.extend([final_audit])
            self.myEssentials.extend([final_essential])
            self.myStandard.extend([final_standard])
            self.myNumber.extend([final_number])
            self.myRequeriment.extend([final_requirement])
            self.myComment.extend([final_comments])
            self.myAuditQuestion.extend([final_question])
            self.myObservation.extend([final_observation])
            self.mySuggested.extend([final_suggested])
            self.myAComments.extend([final_auditcomments])
            self.myPic.extend([final_picture])
        
    def renderExcel(self):
      self.model = TableModel()
      self.table = TableCanvas(self.frame,model=self.model,editable=False,rowheaderwidth=50)
      self.table.createTableFrame()
      self.model = self.table.model

      dict = {}

      for i in range(len(self.myHoja)):
        dict[i] = {'ID': i}

      self.model.importDict(dict)

      self.table.addColumn('Hoja Excel')
      self.table.addColumn('Standard')
      self.table.addColumn('Number')
      self.table.addColumn('Requirement 2015')
      self.table.addColumn('Comments')
      self.table.addColumn('Type of Check')
      self.table.addColumn('Essentials')
      self.table.addColumn('Audit Question')
      self.table.addColumn('Observation / Evidence Required / Audit Remarks')
      self.table.addColumn('Suggested Person to ask')
      self.table.addColumn('Evaluation (0/1)')
      self.table.addColumn('Result')
      self.table.addColumn('Audit Comments')
      # self.table.addColumn('Picture / Statement / Proof')

      for i in range(len(self.myHoja)):
        self.table.model.data[i]['Hoja Excel'] = self.myHoja[i]
        self.table.model.data[i]['Standard'] = self.myStandard[i]
        self.table.model.data[i]['Number'] = self.myNumber[i]
        self.table.model.data[i]['Requirement 2015'] = self.myRequeriment[i]
        self.table.model.data[i]['Comments'] = self.myComment[i]
        self.table.model.data[i]['Type of Check'] = self.myAudit[i]
        self.table.model.data[i]['Essentials'] = self.myEssentials[i]
        self.table.model.data[i]['Audit Question'] = self.myAuditQuestion[i]
        self.table.model.data[i]['Observation / Evidence Required / Audit Remarks'] = self.myObservation[i]
        self.table.model.data[i]['Suggested Person to ask'] = self.mySuggested[i]
        self.table.model.data[i]['Evaluation (0/1)'] = self.myN[i]
        self.table.model.data[i]['Result'] = self.myZero[i]
        self.table.model.data[i]['Audit Comments'] = self.myAComments[i]
        # self.table.model.data[i]['Picture / Statement / Proof'] = self.myPic[i]

      self.table.redrawTable()
    
    #----------------------------------------------------------------------
    """
    LOAD IMAGE Funcionality
    
    1. buttonImage() -> Shows the button to upload the image
    2. loadImage() -> Load the image file
    3. withoutFilter() -> Return a list with cell without None values
    4. loadWorkbook() -> Load from the file, the sheets and each column. Make search functionality
    5. renderExcel() -> Puth the info of file inside of a canvas using tkintertables

    """
    def buttonImage(self):
      self.load_cont_img = Frame(self.image_container)
      self.button_load = Button(self.load_cont_img, justify=LEFT,command=self.loadImage, text="Subir imagen")
      self.button_load.pack()
      self.load_cont_img.pack(side=TOP,fill=X)
    
    def loadImage(self):
      self.file_image = askopenfilename( initialdir = "/KAPTA Camilo/python/img",title = "Subir imagen para guardar en excel", filetypes = (("Imagen de resultado", ".jpg"), ("Todos los archivos", "*.*")))
      if self.file_image is not None:
        self.contImage()
    
    def contImage(self):
      self.rutaImg = Frame(self.image_container)
      Label(self.rutaImg, text="  ").grid(row=1, column=0)
      Label(self.rutaImg, text="  ").grid(row=2, column=0)

      Label(self.rutaImg, text='Ruta de archivo origen', font=("Helvetica", 10), foreground='#E38929').grid(row=1, column=2)
      Label(self.rutaImg, text=self.file_image, font=("Helvetica", 8), foreground='#000').grid(row=2, column=2)

      Label(self.rutaImg, text='Nombre a guardar', font=("Helvetica", 10), foreground='#E38929').grid(row=1, column=4)
      entry = Entry(self.rutaImg,textvariable="entry_txt")
      entry.grid(row=2, column=4)

      Label(self.rutaImg, text='Selecciona Hoja', font=("Helvetica", 10), foreground='#E38929').grid(row=1, column=8)
      ex_sh_sel = StringVar(self.rutaImg)
      ex_sh_sel.set(self.sheets[0])
      w = OptionMenu(self.rutaImg, ex_sh_sel, *self.sheets,command=self.sheet_selected)
      w.grid(row=2,column=8)

      # ex_sh_sel2 = StringVar(self.optionImage)
      # ex_sh_sel2.set('0.0')
      # option2 = OptionMenu(self.optionImage, ex_sh_sel2, '1.0', '2.0', '3.0', '4.0').grid(row=4, column=5)

      # Button(self.optionImage, text="Guardar en Excel", command="saveFile").grid(row=4, column=7)

      # self.optionImage.pack()

      self.rutaImg.pack(side=TOP, fill=X)

    def sheet_selected(self,value):
      item_searched = []

      if(value == self.sheets[0]):
        Label(self.rutaImg, text='Seleccione Item', font=("Helvetica", 10), foreground='#E38929').grid(row=1, column=10)

        self.sheets_search = self.wb.sheetnames[3:4]

        for my_sheet in self.sheets_search:
          my_ws = self.wb[my_sheet]
          for my_row in my_ws.rows:
            if my_row[1].value is not None:
              if not 'Number' in my_row[1].value:
                item_searched.extend([my_row[1].value])
                ex_it_sel = StringVar(self.rutaImg)
                ex_it_sel.set(item_searched[0])
                item = OptionMenu(self.rutaImg, ex_it_sel, *item_searched)
                item.grid(row=2,column=10)

      return value

    def optionsImages(self):
      # self.optionImage = Toplevel(self)
      # self.optionImage.title('K@PTA')
      # self.optionImage.wm_iconbitmap('kapta_mex.ico')
      # self.optionImage.geometry('400x100')
      self.optionImage = Frame(self)

      Label(self.optionImage, text='Selecciona el archivo para guardar', font=("Helvetica", 10)).grid(row=1, column=2)

      # Button(self.optionImage, text="Guardar en Excel", command=loadExceltoSave).grid(row=2, column=2)

      Label(self.optionImage, text='Selecciona Hoja', font=("Helvetica", 10), foreground='#E38929').grid(row=3, column=2)
      ex_sh_sel = StringVar(self.optionImage)
      ex_sh_sel.set('Section 1_Brand Architecture')
      option = OptionMenu(self.optionImage, ex_sh_sel, 'Section 2_OCS', 'Section 4_Customer Area', 'Section 5_IT', 'Section 6_Management','Section 7_Personnel  Training','Section 8_Customer Processes','Section 9_Marketing').grid(row=4, column=2)

      Label(self.optionImage, text='Seleccione Item', font=("Helvetica", 10), foreground='#E38929').grid(row=3, column=5)
      ex_sh_sel2 = StringVar(self.optionImage)
      ex_sh_sel2.set('0.0')
      option2 = OptionMenu(self.optionImage, ex_sh_sel2, '1.0', '2.0', '3.0', '4.0').grid(row=4, column=5)

      Button(self.optionImage, text="Guardar en Excel", command="saveFile").grid(row=4, column=7)

      self.optionImage.pack()
    
    #----------------------------------------------------------------------
    def footer(self):
      self.toolbar = Frame(self,bg='white')
      self.myLabel = Label(self.toolbar, text='Derechos reservados K@PTA', bg='white')
      self.myLabel.pack(side=RIGHT)
      self.toolbar.pack(side=BOTTOM, fill=X)

    #----------------------------------------------------------------------
    def acercaDe(self):
      self.about = Toplevel(self)
      self.about.title('K@PTA')
      # self.about.wm_iconbitmap('kapta_mex.ico')
      self.about.geometry('200x100')
      Label(self.about, text='Derechos Reservados K@PTA \n Desarrollador Camilo Arguello \n Farrell, D 2016 DataExplore: An Application for General Data Analysis in Research and Education. Journal of Open Research Software, 4: e9, DOI: http://dx.doi.org/10.5334/jors.94', font=("Segoe UI", 9), justify=LEFT).pack()
    #----------------------------------------------------------------------
    def onClose(self):
      
      """"""
      self.destroy()
      self.quit()
      root.destroy()
      root.quit()
      exit()
      # self.original_frame.show()
 
########################################################################
class MyApp(object):
    """ Initial PAGE """
 
    #----------------------------------------------------------------------
    def __init__(self, parent):
        """Constructor"""
        self.root = parent

        myW = 600
        myH = 200

        myWs = root.winfo_screenwidth()
        myHs = root.winfo_screenheight()

        myX = (myWs/2) - (myW/2)
        myY = (myHs/2) - (myH/2)


        # self.root.wm_iconbitmap('kapta_mex.ico')
        self.root.geometry('%dx%d+%d+%d' % (myW, myH, myX, myY))
        self.root.title("K@PTA Excel Auditorias")
        self.root.resizable(0,0)
        self.frame = Frame(parent)

        self.label = Label(self.frame, text='Para iniciar por favor carga el archivo de auditorias')
        self.button = Button(self.frame, justify=LEFT,command=self.openFrame)
        self.photo = ImageTk.PhotoImage(file='img/subir_excel.png')
        self.button.config(image=self.photo, width='300',height='50')
        self.label.pack(fill=X)
        self.button.pack()

        self.frame.pack(expand=1)
 
    #----------------------------------------------------------------------
    def hide(self):
        """"""
        self.root.withdraw()
 
    #----------------------------------------------------------------------
    def openFrame(self):
        """"""
        self.hide()
        subFrame = OtherFrame(self)
 
    #----------------------------------------------------------------------
    def show(self):
        """"""
        self.root.update()
        self.root.deiconify()
 

#----------------------------------------------------------------------
if __name__ == "__main__":
  root = Tk()
  app = MyApp(root)
  root.mainloop()
  exit()